#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 13 21:31:23 2017

@author: højlund
"""

# measure how long it takes to run
import time
t0 = time.time()

from collections import OrderedDict
import pandas
import numpy as np
from os.path import join, splitext
from openpyxl import load_workbook

from splinter import Browser
# initialize (headless) browser
browser = Browser('chrome', headless=True) # NB! Chrome extension must be installed for this to work, otherwise just leave blank, and Firefox will be used as default
#browser = Browser(headless=True)
korpus_url = 'http://ordnet.dk/korpusdk_en/concordance'
# open url
browser.visit(korpus_url)

# use pandas to read the excel file, specifying how many rows to skip before reading the header
Pdir = '/Users/au183362/Dropbox/Parkinson_embodied/stimuli/'
file_name = 'Sammenligningsskema.xlsx'
sub_header = 'Comparison content words:' # content of the cell in the first column of the row with the relevant headers
df_dummy = pandas.read_excel(join(Pdir,file_name)) # dummy load of the spreadsheet
first_col = df_dummy[df_dummy.columns[0]].values == sub_header # create boolean for first column
rows_to_skip = [i for i, x in enumerate(first_col) if x] # find True index
#rows_to_skip = 62 # for indholdsord = 62 (NB! doublecheck this), for verbs = 17
df = pandas.read_excel(join(Pdir,file_name),skiprows=rows_to_skip[0] + 1) 

# specifying parameters for reading the excel-flie and saving later
search_terms = ('lemma','word')
searc_pos = ('PoS', 'PoS.1', 'PoS.2', 'PoS.3')
texts = ('Content word', 'Content word.1', 'Content word.2', 'Content word.3')
texts_inf = ('Infinitive', 'Infinitive.1', 'Infinitive.2', 'Infinitive.3')
text_names = ('action', 'non-action', 'act-non-act1', 'act-non-act2')
write_names = (('Frequency_lemma', 'Frequency_word'), ('Frequency_lemma.1', 'Frequency_word.1'), 
               ('Frequency_lemma.2', 'Frequency_word.2'), ('Frequency_lemma.3', 'Frequency_word.3'))
#search_pos = ('PoS', 'PoS.1')
#texts = ('Content word', 'Content word.1')
#texts_inf = ('Infinitive', 'Infinitive.1')
#text_names = ('action', 'non-action')
#write_names = (('Frequency_lemma', 'Frequency_word'), ('Frequency_lemma.1', 'Frequency_word.1'))
word_lists = []
pos_lists = []
occ = OrderedDict()

# loop over the two old texts (aka. action and non-action)
for num, col in enumerate(texts):
    occ[text_names[num]] = OrderedDict()
    word_list_long = df[col].values
    inf_list_long = df[texts_inf[num]].values
    pos_list_long = df[search_pos[num]].values
    nans = np.where(pandas.isnull(word_list_long))[0]
    word_lists.append((inf_list_long[:nans[0]-1]))
    word_lists.append((word_list_long[:nans[0]-1]))
    pos_lists.append((pos_list_long[:nans[0]-1]))
    
    for s, term in enumerate(search_terms):  
        html = []
        occ[text_names[num]][term] = OrderedDict() # initializing OrderedDict in order to keep the order of the dictionary as it was defined
        
        for val, words in enumerate(word_lists[s+num*2]):
            words = words.strip()
            pos = pos_lists[num][val]
                
            browser.find_by_name('formal').click()
            browser.find_by_id('search_box').fill('[' + term + '="' + words + '" & pos="' + pos + '"]')
            browser.find_by_id('search_button').click()
            html.append((browser.html))
            
            reduced = html[val].find('Reduced from ')
            no_result = html[val].find('No results')
            if reduced != -1:
                reduc_end = html[val].find('occurrences',reduced+13)
                occ[text_names[num]][term][words] = int(html[val][reduced+13:reduc_end-1])
            elif no_result != -1:
                occ[text_names[num]][term][words] = 0
            else:
                term_end = html[val].find('occurrences',0)
                term_beg = html[val].rfind('of',0,term_end)
                occ[text_names[num]][term][words] = int(html[val][term_beg+2:term_end])
                
        print(text_names[num])
        print(term)
            
# close (headless) browser
browser.quit()

# save the dictionary as numpy array
np.save('occurrences.npy', occ)

# code needed to load the dictionaries if need be
#occ_lemmas = np.load(join(Pdir,'korpus_dk_search/occ_lemmas.npy')).item()
#occ_words = np.load(join(Pdir,'korpus_dk_search/occ_words.npy')).item()

START_ROW = rows_to_skip+2 # 0 based (subtract 1 from excel row number)
wb = load_workbook(join(Pdir,file_name))
sheet_name = wb.get_sheet_names()[0]
sheet_ranges = wb[sheet_name]
word_list_index = (1,3)

# iterate over the two texts - first lemmas then words
for num, col in enumerate(write_names):
    for s, term in enumerate(search_terms):
        col_write = np.where(df.columns==write_names[num][s])[0][0] # column number for the relevant "Verb" column
        col_read = np.where(df.columns==texts[num])[0][0] # column number for the relevant "Automatic" column
        for row_index in range(START_ROW, sheet_ranges.max_row):
            check_word = sheet_ranges.cell(row=row_index+1, column=col_read+1).value # get the value of the relevant "Verb" cell to match with those in the word_list
            if (row_index-START_ROW) < len(word_lists[word_list_index[num]]):
                if check_word == word_lists[word_list_index[num]][row_index-START_ROW]:
                    # if check_word and word_lists[num][row_index] are the same, then write the frequency in the relevant cell
                    sheet_ranges.cell(row=row_index+1, column=col_write+1, value=occ[text_names[num]][term][word_lists[s+num*2][row_index-START_ROW]])
                    
wb.save(join(Pdir,splitext(file_name)[0]) + '_upd' + splitext(file_name)[-1])

t1 = time.time()
total_time = t1-t0
print(total_time)

